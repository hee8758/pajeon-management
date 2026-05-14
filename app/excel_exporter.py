"""
파견료 내역서 엑셀 출력 모듈
업체별/부서별 청구내역서를 엑셀 파일로 출력합니다.
"""
import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from app.models import Worker, MonthlyPayroll, Company
from app import database as db


def format_number(value):
    """숫자를 콤마 포맷으로 변환"""
    if isinstance(value, (int, float)):
        return f"{int(value):,}"
    return str(value)


def export_billing_statement(year: int, month: int, company_name: str, output_path: str) -> str:
    """
    업체별 파견료 청구내역서 엑셀 출력
    부서별로 시트를 구분하여 출력
    """
    workers = db.get_all_workers()
    payrolls = db.get_monthly_payrolls(year=year, month=month)

    # 해당 업체 근로자 필터
    company_workers = [w for w in workers if w.dispatch_company == company_name and w.status == "재직"]

    if not company_workers:
        return f"'{company_name}' 업체에 재직 중인 근로자가 없습니다."

    # 부서별 그룹핑 (취재2부특집+스포츠 → 취재2부 로 합산)
    DEPT_ORDER = ["취재1부", "취재2부", "편집부"]
    dept_groups = {}
    for w in company_workers:
        dept = w.department or "미분류"
        # 취재2부특집·취재2부스포츠 → 취재2부 로 묶기
        if dept in ("취재2부특집", "취재2부스포츠"):
            key = "취재2부"
        else:
            key = dept
        dept_groups.setdefault(key, []).append(w)

    # 원하는 순서로 정렬
    sorted_depts = sorted(dept_groups.keys(),
                          key=lambda d: DEPT_ORDER.index(d) if d in DEPT_ORDER else 99)

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # 스타일 정의
    header_font  = Font(name='맑은 고딕', bold=True, size=11)
    title_font   = Font(name='맑은 고딕', bold=True, size=14)
    normal_font  = Font(name='맑은 고딕', size=10)
    section_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill  = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 구분 헤더 파란색
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align  = Alignment(horizontal='right',  vertical='center')

    dept_name_map = {
        "취재1부": "영상취재1부 데일리",
        "취재2부": "영상취재2부 (특집/스포츠)",
        "편집부":  "영상편집부 KDNS",
    }
    sub_dept_label = {
        "취재2부특집":   "▶ 영상취재2부 특집",
        "취재2부스포츠": "▶ 영상취재2부 스포츠",
    }

    HEADER_COLS = 24  # A~X

    def write_headers(ws, dept_display, row_offset=0):
        """타이틀·헤더 행 작성, 시작 행 반환"""
        r1 = 1 + row_offset
        ws.merge_cells(f'A{r1}:X{r1}')
        ws[f'A{r1}'] = f'KBS 보도영상 {dept_display} {year}년 {month}월 청구내역서'
        ws[f'A{r1}'].font = title_font
        ws[f'A{r1}'].alignment = center_align

        r2 = 2 + row_offset
        ws.merge_cells(f'A{r2}:X{r2}')
        ws[f'A{r2}'] = f'업체명: {company_name}'
        ws[f'A{r2}'].font = header_font

        h3 = ['순번', '성명', '근무일수', '용역비', '결근일수', '결근공제',
               '대휴 및 야근 공제\n(VAT포함)', '', '용역비\n소 계\n(VAT포함)',
               '시간외수당', '', '', '', '', '', '', '', '', '', '', '', '',
               '합   계\n(VAT포함)', '비   고\n(시간합계)']
        h4 = ['', '', '', '', '', '', '', '', '',
               '평일연장근로수당\n[시급×1.5×1.1]', '',
               '휴일근로수당\n[시급×1.5×1.1]', '',
               '휴일연장근로수당\n[시급×2.0×1.1]', '',
               '야간근로수당\n[시급×0.5×1.1]', '',
               '소  계\n(VAT포함)', '시간외수당\n지급기준', '시간외수당\n실지급액',
               '시간외수당\n간접비', '퇴직급여\n충당금', '', '']
        h5 = ['', '', '', '', '', '', '일수', '금액', '',
               '시간', '금액', '시간', '금액', '시간', '금액', '시간', '금액',
               '', '', '', '', '', '', '']

        for row_data, rn in [(h3, 3 + row_offset),
                              (h4, 4 + row_offset),
                              (h5, 5 + row_offset)]:
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=rn, column=ci, value=val)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

        ws.merge_cells(f'G{3+row_offset}:H{3+row_offset}')
        ws.merge_cells(f'J{3+row_offset}:V{3+row_offset}')
        ws.merge_cells(f'J{4+row_offset}:K{4+row_offset}')
        ws.merge_cells(f'L{4+row_offset}:M{4+row_offset}')
        ws.merge_cells(f'N{4+row_offset}:O{4+row_offset}')
        ws.merge_cells(f'P{4+row_offset}:Q{4+row_offset}')
        return 6 + row_offset  # 데이터 시작 행

    def write_data_rows(ws, workers_list, payrolls, data_start_row):
        """데이터 행 작성, 다음 빈 행 반환"""
        for idx, worker in enumerate(workers_list, 1):
            row = data_start_row + idx - 1
            payroll = next((p for p in payrolls if p.worker_id == worker.id), None)
            ws.cell(row=row, column=1, value=idx).font = normal_font
            ws.cell(row=row, column=2, value=worker.name).font = normal_font
            if payroll:
                vals = [payroll.work_days, payroll.base_fee, payroll.absent_days,
                        payroll.absent_deduction, payroll.leave_deduction_days,
                        payroll.leave_deduction_amount, payroll.fee_subtotal,
                        payroll.weekday_overtime_hours, payroll.weekday_overtime_amount,
                        payroll.holiday_work_hours, payroll.holiday_work_amount,
                        payroll.holiday_overtime_hours, payroll.holiday_overtime_amount,
                        payroll.night_work_hours, payroll.night_work_amount,
                        payroll.overtime_subtotal, payroll.overtime_criteria,
                        payroll.overtime_actual, payroll.overtime_indirect,
                        payroll.retirement_reserve, payroll.total, payroll.total_hours]
            else:
                vals = [0] * 22
            for ci, v in enumerate(vals, 3):
                ws.cell(row=row, column=ci, value=v).font = normal_font
            for col in range(1, 25):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = right_align if col >= 3 else center_align
                if col >= 3:
                    cell.number_format = '#,##0'
        return data_start_row + len(workers_list)

    def write_total_row(ws, data_start, total_row):
        """합계 행 작성"""
        ws.cell(row=total_row, column=2, value='합계').font = header_font
        ws.cell(row=total_row, column=2).alignment = center_align
        for col in range(3, 25):
            sc = f'{get_column_letter(col)}{data_start}'
            ec = f'{get_column_letter(col)}{total_row - 1}'
            c = ws.cell(row=total_row, column=col, value=f'=SUM({sc}:{ec})')
            c.font = header_font
            c.border = thin_border
            c.alignment = right_align
            c.number_format = '#,##0'

    def write_col_widths(ws):
        widths = [5, 8, 7, 12, 7, 10, 5, 10, 12, 6, 12, 6, 12, 6, 12, 6, 12, 12, 12, 12, 12, 12, 14, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 생성 ──
    for dept_key in sorted_depts:
        dept_workers_all = dept_groups[dept_key]
        dept_display = dept_name_map.get(dept_key, dept_key)
        ws = wb.create_sheet(title=dept_key[:31])

        if dept_key == "취재2부":
            # 특집·스포츠 분리
            special  = [w for w in dept_workers_all if w.department == "취재2부특집"]
            sports   = [w for w in dept_workers_all if w.department == "취재2부스포츠"]

            # 헤더 (1회만)
            data_start = write_headers(ws, dept_display, row_offset=0)

            current_row = data_start

            for sub_key, sub_workers in [("취재2부특집", special), ("취재2부스포츠", sports)]:
                if not sub_workers:
                    continue
                # 구분 헤더 행
                label = sub_dept_label.get(sub_key, sub_key)
                ws.merge_cells(f'A{current_row}:X{current_row}')
                c = ws.cell(row=current_row, column=1, value=label)
                c.font = section_font
                c.fill = section_fill
                c.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1

                sub_data_start = current_row
                current_row = write_data_rows(ws, sub_workers, payrolls, sub_data_start)
                # 소계
                write_total_row(ws, sub_data_start, current_row)
                ws.cell(row=current_row, column=2).value = '소계'
                current_row += 1

            # 전체 합계
            ws.cell(row=current_row, column=2, value='총합계').font = header_font
            ws.cell(row=current_row, column=2).alignment = center_align
            for col in range(3, 25):
                cells_ref = []
                # 소계 행들의 합산
                for r in range(data_start, current_row):
                    cell_val = ws.cell(row=r, column=2).value
                    if cell_val in ('소계',):
                        cells_ref.append(f'{get_column_letter(col)}{r}')
                if cells_ref:
                    formula = '=' + '+'.join(cells_ref)
                else:
                    formula = 0
                c2 = ws.cell(row=current_row, column=col, value=formula)
                c2.font = header_font
                c2.border = thin_border
                c2.alignment = right_align
                c2.number_format = '#,##0'
            current_row += 1

            # 공급가액/부가세
            ws.cell(row=current_row, column=1, value='공급가액').font = header_font
            ws.cell(row=current_row, column=9, value='부가세').font = header_font
            ws.cell(row=current_row, column=12, value='합  계').font = header_font

        else:
            # 일반 부서 (기존 방식)
            data_start = write_headers(ws, dept_display, row_offset=0)
            total_row  = write_data_rows(ws, dept_workers_all, payrolls, data_start)
            write_total_row(ws, data_start, total_row)
            tax_row = total_row + 1
            ws.cell(row=tax_row, column=1, value='공급가액').font = header_font
            ws.cell(row=tax_row, column=9, value='부가세').font  = header_font
            ws.cell(row=tax_row, column=12, value='합  계').font = header_font

        write_col_widths(ws)

    try:
        wb.save(output_path)
        return f"청구내역서가 저장되었습니다: {output_path}"
    except Exception as e:
        return f"저장 실패: {str(e)}"


def export_all_companies(year: int, month: int, output_dir: str) -> str:
    """모든 업체의 청구내역서를 한 번에 출력"""
    companies = db.get_all_companies()
    messages = []

    for company in companies:
        filename = f"파견료내역서_{company.name}_{year}년{month}월.xlsx"
        output_path = os.path.join(output_dir, filename)
        msg = export_billing_statement(year, month, company.name, output_path)
        messages.append(msg)

    return "\n".join(messages) if messages else "출력할 업체가 없습니다."
